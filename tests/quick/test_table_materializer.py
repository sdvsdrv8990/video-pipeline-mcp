"""
tests/quick/test_table_materializer.py — A1′: материализация книг по .schema.yaml (фаза ТАБЛИЦЫ).

Standalone-прогон:  python tests/quick/test_table_materializer.py
Проверяет: e2e proof-схема network_config → .xlsx (листы/столбцы/enum-валидация/формулы),
контракт ошибок (нет схемы, битая схема, книга уже есть), фаза по tables_pending
(отказ одной книги не роняет соседние).
"""
import sys
import tempfile
import warnings
from pathlib import Path

import openpyxl

warnings.simplefilter("error", UserWarning)  # чужой код ошибки (D25/G14) → падение

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from core.engine import TableMaterializer, TableMaterializerError
from core.excel import ExcelError
from core.excel import ExcelEngine

SCHEMAS = ROOT / "config" / "templates" / "tables"

_checks = 0
_fails = []


def ok(cond, msg):
    global _checks
    _checks += 1
    if not cond:
        _fails.append(msg)
        print(f"  ✗ {msg}")
    else:
        print(f"  ✓ {msg}")


def new_materializer(ws: str):
    return TableMaterializer(ExcelEngine(ws), SCHEMAS)


ws = tempfile.mkdtemp(prefix="tm_")
mat = new_materializer(ws)

print("== 1. e2e: proof-схема network_config → книга ==")
res = mat.materialize("network_config", "networks/net_1/network_config.xlsx")
book = Path(ws) / "networks" / "net_1" / "network_config.xlsx"
ok(book.exists(), "книга материализована на диске")
ok(res["book"] == "network_config" and res["level"] == "network", "в результате книга и уровень из схемы")
ok(len(res["sheets"]) == 4, f"4 листа из схемы (получено {len(res['sheets'])})")
ok(res["columns_total"] == 17, f"17 столбцов суммарно (5+5+3+4 по схеме) (получено {res['columns_total']})")

print("== 2. Форма книги в .xlsx соответствует декларации ==")
wb = openpyxl.load_workbook(book)
ok(wb.sheetnames == ["NETWORK_SCHEDULE_MASTER", "SHARED_RESOURCES", "CROSS_CHANNEL_RULES", "NETWORK_GATES"],
   f"порядок листов как в схеме: {wb.sheetnames}")
headers = [c.value for c in wb["NETWORK_SCHEDULE_MASTER"][1]]
ok(headers == ["date", "channel_id", "planned_topic", "status", "conflict_check"],
   f"заголовки первого листа: {headers}")
ok(len(wb["SHARED_RESOURCES"].data_validations.dataValidation) == 0,
   "лист без enum-колонок не получил лишних валидаций")

print("== 3. type: enum → выпадающий список (set_validation) ==")
dvs = wb["NETWORK_SCHEDULE_MASTER"].data_validations.dataValidation
ok(len(dvs) == 1, f"одна валидация на лист с одним enum (получено {len(dvs)})")
ok(dvs and "PLANNED" in dvs[0].formula1 and "SCHEDULED" in dvs[0].formula1,
   "значения enum из схемы попали в дропдаун")
dvs_rules = wb["CROSS_CHANNEL_RULES"].data_validations.dataValidation
ok(dvs_rules and "HARD" in dvs_rules[0].formula1, "enum второго листа (HARD/SOFT) тоже материализован")

print("== 4. Ошибки идут кодом реестра, не текстом ==")
try:
    mat.materialize("нет_такой_книги", "x.xlsx")
    ok(False, "несуществующая схема должна падать")
except TableMaterializerError as e:
    ok(e.code == "TEMPLATE_NOT_FOUND", f"нет схемы → TEMPLATE_NOT_FOUND (получено {e.code})")
    ok(bool(e.reason), "в ошибке есть actionable reason")

bad = SCHEMAS / "_tmp_broken.schema.yaml"
bad.write_text("book: broken\nsheets: []\n", encoding="utf-8")
try:
    mat.materialize("_tmp_broken", "broken.xlsx")
    ok(False, "схема без листов должна падать")
except TableMaterializerError as e:
    ok(e.code == "SCHEMA_INVALID", f"схема без листов → SCHEMA_INVALID (получено {e.code})")
finally:
    bad.unlink()

try:
    mat.materialize("network_config", "networks/net_1/network_config.xlsx")
    ok(False, "существующая книга не должна молча перезаписываться")
except TableMaterializerError as e:
    ok(e.code == "FILE_EXISTS", f"книга уже есть → FILE_EXISTS (получено {e.code})")

print("== 5. Фаза ТАБЛИЦЫ: обход tables_pending от structure_create ==")
ws2 = tempfile.mkdtemp(prefix="tm2_")
mat2 = new_materializer(ws2)
pending = [
    {"path": "networks/n1/network_config.xlsx", "table_template": "network_config",
     "required": True, "file_id": "FILE_1"},
    {"path": "networks/n2/network_config.xlsx", "table_template": "network_config",
     "required": True, "file_id": "FILE_2"},
    {"path": "networks/n3/unknown.xlsx", "table_template": "ещё_не_заведена",
     "required": False, "file_id": "FILE_3"},
    {"path": "networks/n4/nameless.xlsx", "required": False, "file_id": "FILE_4"},
]
phase = mat2.materialize_pending(pending)
ok(phase["created"] == 2 and phase["total"] == 4, f"2 из 4 книг созданы: {phase['created']}/{phase['total']}")
ok(len(phase["failed"]) == 2, "две записи отчитались отказом, а не исключением")
ok({f["code"] for f in phase["failed"]} == {"TEMPLATE_NOT_FOUND", "SCHEMA_INVALID"},
   f"коды отказов из реестра: {[f['code'] for f in phase['failed']]}")
ok(all(m["file_id"] for m in phase["materialized"]), "file_id из tables_pending доехал до результата")
ok((Path(ws2) / "networks" / "n2" / "network_config.xlsx").exists(),
   "отказ соседней книги не помешал материализации следующей")

print("== 6. Контракт с structure_create: pending РЕАЛЬНОГО движка шаблонов ==")
from core.engine import TemplateEngine
from core.ids import IDGenerator

ws3 = tempfile.mkdtemp(prefix="tm3_")
tpl = TemplateEngine(ws3, IDGenerator(), ROOT / "config" / "templates" / "workspace")
node = tpl.create_node("network", "net_A")
pending_real = node["tables_pending"]
ok(len(pending_real) >= 2, f"structure_create отложил книги сетки: {len(pending_real)}")
ok(any(p["table_template"] == "network_config" for p in pending_real),
   "среди отложенного есть network_config")

phase_real = TableMaterializer(ExcelEngine(ws3), SCHEMAS).materialize_pending(pending_real)
# Ожидание выводим из диска: заведённых схем становится больше по мере авторинга.
_with_schema = sum(1 for p in pending_real if (SCHEMAS / f"{p['table_template']}.schema.yaml").exists())
ok(phase_real["created"] == _with_schema,
   f"материализованы ровно книги с заведённой схемой ({phase_real['created']} из {len(pending_real)})")
ok(phase_real["materialized"] and (Path(ws3) / phase_real["materialized"][0]["path"]).exists(),
   "книга легла ровно по пути из tables_pending")
ok(all(f["code"] == "TEMPLATE_NOT_FOUND" for f in phase_real["failed"]),
   "книги без схемы честно отчитались TEMPLATE_NOT_FOUND (G16: не молчаливый успех)")

print("== 7. Директива владельца: клиент передаёт ИМЕНА, книги делает сервер ==")
import asyncio
from core.engine import Engine
from core.ids import LinkRegistry
from core.reactions import Reactions
from core.state import StateManager
from tools._context import ToolContext
from tools import structure as structure_group

ws4 = Path(tempfile.mkdtemp(prefix="tm4_"))
CFG = ROOT / "config"
sm = StateManager(ws4)
ids = IDGenerator()
eng = Engine(reactions=Reactions(CFG / "server_reactions.yaml"), state_manager=sm)
ctx = ToolContext(eng, ids, sm, None, ExcelEngine(ws4),
                  TemplateEngine(ws4, ids, CFG / "templates" / "workspace"),
                  LinkRegistry(ws4), ws4, CFG)
structure_group.register(eng, ctx)

res = asyncio.run(eng.tools["structure_create"].handler(
    type="network", name="fit_net",
    children={"channel": ["ch_A"], "competitor_channel": [f"comp_{i}" for i in range(20)]}))
data = res.data
ok(res.status == "success", "пакетное создание (сетка + канал + 20 конкурентов) одним вызовом")
ok(len(list(ws4.rglob("*.xlsx"))) == len(data["tables_materialized"]),
   "книг на диске ровно столько, сколько отчитано материализованными")
ok(data["tables_materialized"] and data["tables_materialized"][0]["book"] == "network_config",
   "книга создана у созданной сущности (сетка), без участия клиента")
ok(not list(ws4.rglob("videos/*/")),
   "видео не названы → сущностей видео нет, книг видео тоже нет (умное создание)")
ok(all(f["code"] == "TEMPLATE_NOT_FOUND" for f in data["tables_deferred"]),
   "книги без декларации отложены честно, а не «созданы» (G16)")
ok(any(f.type == "TableMaterialized" for f in res.facts),
   "факт TableMaterialized доехал в контракт (тип заведён в KNOWN_FACT_TYPES, D25)")


print("== 7b. Строки-дефолты из схемы доезжают В КНИГУ, а не только в декларацию ==")
ws5 = tempfile.mkdtemp(prefix="tm5_")
r5 = new_materializer(ws5).materialize("channel_data", "ch.xlsx")
wb5 = openpyxl.load_workbook(Path(ws5) / "ch.xlsx")
# Считаем по самой схеме: новый провайдер добавляет строки (S24 — фон и апскейл), и сверка с
# числом, вписанным сюда однажды, ловила бы не потерю дефолтов, а факт их добавления.
import yaml
_declared_rows = sum(len(s.get("rows") or []) for s in yaml.safe_load(
    (ROOT / "config/templates/tables/channel_data.schema.yaml").read_text(encoding="utf-8"))["sheets"])
ok(r5["rows_total"] == _declared_rows,
   f"все строки-дефолты схемы доехали в книгу ({r5['rows_total']} из {_declared_rows})")
_sp = wb5["SCENE_PROFILE"]
_vals = {row[0]: row[1] for row in _sp.iter_rows(min_row=2, values_only=True) if row[0]}
ok(len(_vals) == 7, f"SCENE_PROFILE: 7 строк данных на листе (получено {len(_vals)})")
ok(_vals.get("sound") is False and _vals.get("svg_bg") is True,
   "значения дефолтов легли как есть: выключенный тип остался выключенным (тихий столбец)")
_wf = wb5["WORKFLOW_SEQUENCES"]
_hdr = [c.value for c in _wf[1]]
_first = next(_wf.iter_rows(min_row=2, values_only=True), ())     # строк может не быть вовсе
_row2 = {_hdr[i]: v for i, v in enumerate(_first)}
ok(_row2.get("allowed_next_tools") == "prepare_tts_input,trigger_tts_generation,human_approve",
   "список в ячейке — через запятую, значения не потеряны")
ok(isinstance(_row2.get("requires_human_approval"), bool),
   "булево осталось булевым, а не строкой 'False'")

print("== 7c. F30: формулы деградируют на неполных данных, а не ломаются ==")
import shutil as _shutil2
import yaml as _y2

_sch30 = Path(tempfile.mkdtemp(prefix="deg_"))
_shutil2.copy(SCHEMAS / "_defaults.yaml", _sch30 / "_defaults.yaml")
(_sch30 / "probe.schema.yaml").write_text(_y2.safe_dump({
    "book": "probe",
    "sheets": [{"name": "CALC", "columns": [
        {"name": "views", "type": "integer", "flag": "W"},
        {"name": "likes", "type": "integer", "flag": "W"},
        {"name": "like_rate", "type": "float", "flag": "F", "formula": "=B2/A2*100"},
        {"name": "verdict", "type": "string", "flag": "F", "formula": "=B2/A2"},
        {"name": "note", "type": "string", "flag": "F", "formula": "=B2/A2", "on_empty": "НЕТ ДАННЫХ"},
    ]}]}, allow_unicode=True), encoding="utf-8")
_ws30 = tempfile.mkdtemp(prefix="deg_ws_")
_e30 = ExcelEngine(_ws30)
_r30 = TableMaterializer(_e30, _sch30).materialize("probe", "p.xlsx")
ok(_r30["formulas_guarded"] == 3, f"объявленные формулы обёрнуты (получено {_r30['formulas_guarded']})")
_f30 = [c.value for c in openpyxl.load_workbook(Path(_ws30) / "p.xlsx")["CALC"][2]]
ok(all(str(v).startswith("=IFERROR(") for v in _f30[2:]), f"в книгу легла защищённая формула: {_f30[2]}")

# Пересчёт РЕАЛЬНЫМ движком: без него это была бы проверка строки, а не поведения (урок F29).
try:
    _calc30 = _e30._recalc_via_lo(Path(_ws30) / "p.xlsx")
    _vals30 = [c.value for c in openpyxl.load_workbook(_calc30, data_only=True)["CALC"][2]]
    ok(_vals30[2] == 0, f"деление на пустое → 0 для float, а не #DIV/0! (получено {_vals30[2]!r})")
    ok(_vals30[3] == "PENDING", f"строковый столбец → PENDING (получено {_vals30[3]!r})")
    ok(_vals30[4] == "НЕТ ДАННЫХ", f"on_empty столбца перекрывает правило типа (получено {_vals30[4]!r})")
    # Контроль: та же формула БЕЗ деградации обязана дать ошибку — иначе проверка холостая.
    _wsc = tempfile.mkdtemp(prefix="deg_ctl_")
    _ec = ExcelEngine(_wsc)
    _ec.create_workbook("c.xlsx", sheet="CALC")
    for _c in ("views", "likes"):
        _ec.add_column("c.xlsx", "CALC", _c)
    _ec.add_column("c.xlsx", "CALC", "like_rate", formula="=B2/A2*100")
    _ctl = [c.value for c in openpyxl.load_workbook(_ec._recalc_via_lo(Path(_wsc) / "c.xlsx"),
                                                    data_only=True)["CALC"][2]]
    ok(_ctl[2] == "#DIV/0!", f"контроль: без деградации та же формула даёт #DIV/0! (получено {_ctl[2]!r})")
except ExcelError as _e:
    print(f"    (пересчёт недоступен: {_e.code} — проверки значений пропущены)")

# Правила деградации — декларация, а не код: в движке нет ни одного зашитого запасного значения.
_src30 = (ROOT / "core/engine/table_materializer.py").read_text(encoding="utf-8")
ok("PENDING" not in _src30, "запасные значения не зашиты в код — только читаются из _defaults.yaml")

print("== 8. Конвертер спека→схема: что разобрано, а что честно отдано человеку ==")
import importlib.util

_spec = importlib.util.spec_from_file_location("s2s", ROOT / "scripts" / "spec_to_schema.py")
s2s = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(s2s)

ok(s2s.verify() == 0, "разбор сходится с собранной РУКАМИ network_config (приёмка конвертера)")

specs = Path(tempfile.mkdtemp(prefix="s2s_"))


def parse(body: str) -> dict:
    p = specs / "probe.schema.md"
    p.write_text(body, encoding="utf-8")
    return s2s.parse_spec(p)


def cols_of(res: dict, i: int = 0) -> list[dict]:
    """Столбцы листа, не роняя прогон: при мутации листа может не быть вовсе."""
    return res["sheets"][i]["columns"] if len(res["sheets"]) > i else []


# Диапазон в ДВУХ парах кавычек раньше проходил немо: имя бралось из первой пары,
# многоточие оставалось в хвосте заголовка, и 7 листов схлопывались в один без слова.
r = parse("## Листы 4–10: `ACT_1_HOOK` … `ACT_7_TRUTH` (7 идентичных)\n\n"
          "`act_name, text` — `W`.\n")
ok(not r["sheets"], "диапазон листов не превращается в один лист")
ok(any("не называет их поимённо" in w for w in r["warnings"]),
   "диапазон без поимённого перечисления — предупреждение, а не догадка об именах")

# Составное имя разворачивается ТОЛЬКО потому, что спека сама называет листы в «Дельтах».
r = parse("## Листы 5–6: `VISUAL_/SCRIPT_LIB` (шаблон)\n\n"
          "**Общие:** `solution_id (id), notes` — всё `W`.\n\n"
          "**Дельты:**\n"
          "- `VISUAL_LIB` (5): + `visual_solution`.\n"
          "- `SCRIPT_LIB` (6): + `pattern_description`.\n")
ok([s["name"] for s in r["sheets"]] == ["VISUAL_LIB", "SCRIPT_LIB"],
   "группа разворачивается в листы, ПОИМЕННО названные спекой")
ok([c["name"] for c in cols_of(r)] == ["solution_id", "notes", "visual_solution"],
   "у листа группы общие столбцы + собственная дельта")
ok(bool(cols_of(r)) and cols_of(r)[0]["flag"] == "id", "флаг из аннотации `(id)` доехал")

# Прозаическая форма: имена и флаги спека объявляет, тип — нет.
r = parse("## Лист 1: `META`\n\n"
          "`video_id (id), title, channel_id (fk), tier (HIGH/MED/LOW)` — `W`;\n"
          "`type` = `F` (фикс.).\n")
cols = {c["name"]: c for c in cols_of(r)}
ok(list(cols) == ["video_id", "title", "channel_id", "tier", "type"],
   "прозаический список даёт столбцы, включая одиночный `type` = `F` вне списка")
ok(cols.get("channel_id", {}).get("flag") == "fk" and cols.get("type", {}).get("flag") == "F",
   "флаги из прозы верны")
ok(cols.get("tier", {}).get("type") == "enum" and cols.get("tier", {}).get("enum") == ["HIGH", "MED", "LOW"],
   "enum-значения из скобок, а не выдуманные")
ok(bool(r["sheets"]) and r["sheets"][0]["prose"], "лист помечен как собранный из прозы (тип не объявлен спекой)")

# Не выдумываем: составной столбец `a_x/y` — это ДВА столбца, развести может только человек.
r = parse("## Лист 1: `S`\n\n`scene_id (id), color_primary/secondary, notes` — `W`.\n")
names = [c["name"] for c in cols_of(r)]
ok("color_primary/secondary" not in names and "color_primary" not in names,
   "составной столбец через `/` не разводится машиной и не попадает в схему как есть")
ok(any("не приняты за имена столбцов" in w for w in r["warnings"]),
   "отброшенный токен назван поимённо — это адрес для вычитки")

# Дашборд из секций собирается, но полнота набора под вопросом → сервер обязан сказать.
r = parse("## Лист 9: `ANALYTICS` — дашборд (Read-Only)\n\n"
          "> Весь `F`. Секции: **Топ** (`metric, delta`) · **Прочее** (trend, our_status).\n")
ok([c["flag"] for c in cols_of(r)] == ["F", "F"],
   "лист, объявленный целиком Read-Only, даёт столбцы с флагом F")
ok(any("проверить полноту" in w for w in r["warnings"]),
   "секционный дашборд помечен как возможно неполный, а не выдан за полный")


print(f"\n{'='*50}")
print(f"РЕЗУЛЬТАТ: {_checks - len(_fails)}/{_checks} прошло")
if _fails:
    print("ПРОВАЛЫ:")
    for f in _fails:
        print(f"  - {f}")
    sys.exit(1)
print("ВСЁ ЗЕЛЁНОЕ ✅")
