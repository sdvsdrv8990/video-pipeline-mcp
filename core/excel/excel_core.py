"""core/excel/excel_core.py — движок СТРУКТУРЫ таблиц (Excel Engine).

## Назначение
CRUD над ФОРМОЙ книги: листы, столбцы, формулы, форматирование, валидация. Значения ячеек
здесь не пишутся — их меняет очередь слоя данных (`core/tables`).

## Границы
Заголовки столбцов — строка 1, данные со строки 2; столбец адресуется ПО ИМЕНИ заголовка,
а не по букве. `read_range` — отладочное сырое чтение, рабочее идёт через снапшот и проекции.
`insert_formula` не перезаписывает существующую формулу молча (FORMULA_PROTECTED).
Все пути — через `core.paths.safe_resolve` внутри `workspace/` (G17/D29).
"""

from __future__ import annotations

import re
from pathlib import Path

from core.paths import safe_resolve

# Ссылка в формуле: необязательный лист (`META!` / `'Мой лист'!`), затем ячейка или
# диапазон (`B2`, `$B$2`, `A2:D9`). Границы обязательны: без них `LOG10(` в `=LOG10(x)`
# прочиталось бы как ячейка LOG10 и дало ложную зависимость (F28).
CELL_REF_RE = re.compile(
    r"(?<![A-Za-z0-9_$!])"
    r"(?:(?:'(?P<quoted>[^']+)'|(?P<plain>[A-Za-z_][A-Za-z0-9_.]*))!)?"
    r"\$?(?P<c1>[A-Z]{1,3})\$?(?P<r1>\d+)"
    r"(?::\$?(?P<c2>[A-Z]{1,3})\$?(?P<r2>\d+))?"
    r"(?![A-Za-z0-9_(])"
)

# openpyxl импортируем лениво внутри методов — чтобы импорт модуля не падал,
# если библиотека не установлена (движок данных Категории 3 от неё не зависит).


class ExcelError(Exception):
    """Ошибка движка структуры. Код из server_reactions.yaml + подсказка."""

    def __init__(self, code: str, message: str, reason: str = "", suggested_tool: str | None = None):
        super().__init__(message)
        self.code = code
        self.message = message
        self.reason = reason
        self.suggested_tool = suggested_tool


class ExcelEngine:
    """Движок структуры Excel-книг. Один на категорию.

    Attributes:
        workspace: корень workspace/ для containment.
    """

    def __init__(self, workspace: str | Path):
        self.workspace = Path(workspace)

    # ═══ Служебное ═══

    def _resolve(self, path: str) -> Path:
        """safe-join внутри workspace. ValueError → обёртка маппит в PATH_ESCAPE."""
        return safe_resolve(path, self.workspace)

    def _load(self, path: str, must_exist: bool = True):
        """Открыть книгу. WORKBOOK_NOT_FOUND если нет (и must_exist)."""
        import openpyxl
        p = self._resolve(path)
        if not p.exists():
            if must_exist:
                raise ExcelError(
                    "WORKBOOK_NOT_FOUND", f"Книга не найдена: {path}",
                    reason="Создай книгу через excel_create_workbook или проверь путь.",
                    suggested_tool="excel_create_workbook",
                )
            return None
        return openpyxl.load_workbook(p)

    def _sheet(self, wb, sheet: str):
        if sheet not in wb.sheetnames:
            raise ExcelError(
                "SHEET_NOT_FOUND", f"Лист '{sheet}' не найден. Есть: {', '.join(wb.sheetnames)}",
                reason="Проверь имя листа или создай через excel_add_sheet.",
            )
        return wb[sheet]

    @staticmethod
    def _headers(ws) -> dict[str, int]:
        """{имя_заголовка: индекс_столбца(1-based)} из строки 1."""
        headers: dict[str, int] = {}
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value is not None:
                headers[str(cell.value)] = idx
        return headers

    def _save(self, wb, path: str):
        wb.save(self._resolve(path))

    # ═══ КНИГА / ЛИСТЫ ═══

    def create_workbook(self, path: str, sheet: str = "Sheet1") -> dict:
        """Новый .xlsx. FILE_EXISTS если файл уже есть (не перезаписываем молча)."""
        import openpyxl
        p = self._resolve(path)
        if p.exists():
            raise ExcelError("FILE_EXISTS", f"Файл или каталог уже существует: {path}",
                             reason="Удали существующий файл или используй другое имя.")
        p.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.active.title = sheet
        wb.save(p)
        return {"path": path, "sheet": sheet}

    def add_sheet(self, path: str, sheet: str) -> dict:
        wb = self._load(path)
        if sheet in wb.sheetnames:
            raise ExcelError("SHEET_EXISTS", f"Лист уже существует: {sheet}",
                             reason="Используй другое имя или excel_rename_sheet.")
        wb.create_sheet(title=sheet)
        self._save(wb, path)
        return {"path": path, "sheet": sheet, "sheets": wb.sheetnames}

    def rename_sheet(self, path: str, sheet: str, new_name: str) -> dict:
        wb = self._load(path)
        ws = self._sheet(wb, sheet)
        if new_name in wb.sheetnames:
            raise ExcelError("SHEET_EXISTS", f"Лист уже существует: {new_name}",
                             reason="Выбери свободное имя.")
        ws.title = new_name
        self._save(wb, path)
        return {"path": path, "renamed": {sheet: new_name}, "sheets": wb.sheetnames}

    def delete_sheet(self, path: str, sheet: str) -> dict:
        wb = self._load(path)
        self._sheet(wb, sheet)
        if len(wb.sheetnames) == 1:
            raise ExcelError("LAST_SHEET", "Нельзя удалить последний лист книги.",
                             reason="В книге всегда должен быть хотя бы один лист.")
        del wb[sheet]
        self._save(wb, path)
        return {"path": path, "deleted": sheet, "sheets": wb.sheetnames}

    def reorder_sheets(self, path: str, order: list[str]) -> dict:
        wb = self._load(path)
        if set(order) != set(wb.sheetnames):
            raise ExcelError("VALIDATION_ERROR", "order должен содержать РОВНО все листы книги.",
                             reason=f"Ожидались листы: {wb.sheetnames}.")
        wb._sheets.sort(key=lambda ws: order.index(ws.title))
        self._save(wb, path)
        return {"path": path, "sheets": wb.sheetnames}

    def copy_sheet(self, path: str, sheet: str, new_name: str) -> dict:
        """Копирование листа с данными и форматированием."""
        wb = self._load(path)
        self._sheet(wb, sheet)
        if new_name in wb.sheetnames:
            raise ExcelError("SHEET_EXISTS", f"Лист уже существует: {new_name}",
                             reason="Выбери свободное имя.")
        source = wb[sheet]
        copy = wb.copy_worksheet(source)
        copy.title = new_name
        self._save(wb, path)
        return {"path": path, "copied": sheet, "to": new_name, "sheets": wb.sheetnames}

    # ═══ АНАЛИЗ СТРУКТУРЫ ═══

    def inspect_file(self, path: str) -> dict:
        """Обзор структуры книги: листы, размеры, формат."""
        wb = self._load(path)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append({
                "name": name,
                "rows": ws.max_row,
                "columns": ws.max_column,
            })
        return {
            "path": path,
            "format": self._resolve(path).suffix,
            "sheet_count": len(wb.sheetnames),
            "sheets": sheets,
        }

    def get_sheet_info(self, path: str, sheet: str) -> dict:
        """Детальный анализ листа: колонки, типы, превью."""
        wb = self._load(path)
        ws = self._sheet(wb, sheet)
        headers = self._headers(ws)
        columns = []
        for name, idx in headers.items():
            col_type = "string"
            sample_values = []
            for row in range(2, min(ws.max_row + 1, 7)):
                val = ws.cell(row=row, column=idx).value
                if val is not None:
                    sample_values.append(val)
                    if isinstance(val, (int, float)):
                        col_type = "number"
                    elif isinstance(val, bool):
                        col_type = "bool"
            columns.append({
                "name": name,
                "index": idx,
                "type": col_type,
                "sample": sample_values[:3],
            })
        return {
            "path": path,
            "sheet": sheet,
            "row_count": ws.max_row - 1,
            "column_count": len(headers),
            "columns": columns,
        }

    def get_column_names(self, path: str, sheet: str) -> dict:
        """Быстрый список колонок листа."""
        wb = self._load(path)
        ws = self._sheet(wb, sheet)
        headers = self._headers(ws)
        return {
            "path": path,
            "sheet": sheet,
            "columns": list(headers.keys()),
            "count": len(headers),
        }

    # ═══ СТОЛБЦЫ ═══

    def add_column(self, path: str, sheet: str, column: str, formula: str | None = None) -> dict:
        """Новый столбец = заголовок в строку 1 (следующий свободный)."""
        wb = self._load(path)
        ws = self._sheet(wb, sheet)
        headers = self._headers(ws)
        if column in headers:
            raise ExcelError("COLUMN_EXISTS", f"Столбец уже существует: {column}",
                             reason="Используй другое имя или excel_delete_column.")
        new_idx = (max(headers.values()) + 1) if headers else 1
        ws.cell(row=1, column=new_idx, value=column)
        if formula:
            # формула-образец в строку 2 (шаблон вычисляемого столбца)
            ws.cell(row=2, column=new_idx, value=formula if formula.startswith("=") else f"={formula}")
        self._save(wb, path)
        return {"path": path, "sheet": sheet, "column": column, "index": new_idx}

    def append_row(self, path: str, sheet: str, values: dict) -> dict:
        """Строка данных в конец листа: {заголовок: значение}.

        Заголовок обязан существовать — молча создавать столбец под неизвестный ключ нельзя,
        иначе опечатка в декларации тихо расширяет книгу.
        """
        wb = self._load(path)
        ws = self._sheet(wb, sheet)
        headers = self._headers(ws)
        unknown = [k for k in values if k not in headers]
        if unknown:
            raise ExcelError("COLUMN_NOT_FOUND",
                             f"Лист '{sheet}': нет столбцов {unknown}.",
                             reason="Сверь имена с заголовками (excel_get_column_names).")
        # Пишем после последней непустой строки: у листа с формулой-образцом занята строка 2.
        last = max((c.row for row in ws.iter_rows() for c in row if c.value is not None), default=1)
        for name, value in values.items():
            ws.cell(row=last + 1, column=headers[name], value=value)
        self._save(wb, path)
        return {"path": path, "sheet": sheet, "row": last + 1, "written": len(values)}

    def find_dependents(self, path: str, sheet: str, column: str) -> dict:
        """Формулы, ссылающиеся на столбец (F28). Читающая операция — ничего не меняет.

        Ищем по ВСЕЙ книге: ссылка бывает межлистовой (`META!B2`). Диапазон `A2:D9` считается
        ссылкой на каждый столбец внутри него — удаление любого из них ломает диапазон.
        """
        wb = self._load(path)
        ws = self._sheet(wb, sheet)
        headers = self._headers(ws)
        if column not in headers:
            raise ExcelError("COLUMN_NOT_FOUND", f"Столбец '{column}' не найден в листе '{sheet}'.",
                             reason="Сверь имя заголовка (excel_get_column_names).")
        deps = self._dependents(wb, sheet, headers[column])
        return {"path": path, "sheet": sheet, "column": column,
                "dependents": deps, "count": len(deps)}

    @staticmethod
    def _dependents(wb, sheet: str, col_idx: int) -> list[dict]:
        """Ячейки-формулы, чьи ссылки накрывают столбец col_idx листа sheet."""
        from openpyxl.utils import column_index_from_string
        found = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str) or not value.startswith("="):
                        continue
                    for m in CELL_REF_RE.finditer(value):
                        # Лист ссылки: явный префикс либо лист самой формулы.
                        ref_sheet = m.group("quoted") or m.group("plain") or ws.title
                        if ref_sheet != sheet:
                            continue
                        try:
                            first = column_index_from_string(m.group("c1"))
                            last = column_index_from_string(m.group("c2") or m.group("c1"))
                        except ValueError:
                            continue                       # не ссылка (напр. имя функции)
                        if min(first, last) <= col_idx <= max(first, last):
                            found.append({"sheet": ws.title, "cell": cell.coordinate,
                                          "formula": value, "ref": m.group(0)})
                            break                          # одна ячейка — одна запись
        return found

    def _guard_column(self, wb, sheet: str, column: str, col_idx: int,
                      action: str, force: bool) -> list[dict]:
        """Пред-проверка деструктива над столбцом: молчаливый коррапт формул запрещён (F28)."""
        deps = self._dependents(wb, sheet, col_idx)
        if deps and not force:
            where = ", ".join(f"{d['sheet']}!{d['cell']}" for d in deps[:5])
            raise ExcelError(
                "COLUMN_HAS_DEPENDENTS",
                f"На столбец '{column}' ссылаются формулы ({len(deps)}): {where}"
                + (" …" if len(deps) > 5 else ""),
                reason=f"{action} сломает эти формулы (#REF!). Перепиши их или передай force=true осознанно.")
        return deps

    def delete_column(self, path: str, sheet: str, column: str, force: bool = False) -> dict:
        wb = self._load(path)
        ws = self._sheet(wb, sheet)
        headers = self._headers(ws)
        if column not in headers:
            raise ExcelError("COLUMN_NOT_FOUND", f"Столбец '{column}' не найден в листе '{sheet}'.",
                             reason="Сверь имя заголовка (excel_read_range) или уже удалён.")
        broken = self._guard_column(wb, sheet, column, headers[column], "Удаление", force)
        ws.delete_cols(headers[column], 1)
        self._save(wb, path)
        return {"path": path, "sheet": sheet, "deleted": column,
                "broken_formulas": broken}          # непусто только при force=true

    def move_column(self, path: str, sheet: str, column: str, to_index: int, force: bool = False) -> dict:
        """Переместить столбец на позицию to_index (1-based) сдвигом ячеек."""
        wb = self._load(path)
        ws = self._sheet(wb, sheet)
        headers = self._headers(ws)
        if column not in headers:
            raise ExcelError("COLUMN_NOT_FOUND", f"Столбец '{column}' не найден в листе '{sheet}'.",
                             reason="Сверь имя заголовка.")
        from_index = headers[column]
        ncols = max(headers.values())
        if not (1 <= to_index <= ncols):
            raise ExcelError("VALIDATION_ERROR", f"to_index вне диапазона 1..{ncols}.",
                             reason="Укажи позицию внутри существующих столбцов.")
        if to_index == from_index:
            return {"path": path, "sheet": sheet, "column": column, "index": to_index,
                    "broken_formulas": []}
        # Перенос меняет БУКВУ столбца: формулы продолжат указывать на старую позицию.
        broken = self._guard_column(wb, sheet, column, from_index, "Перенос", force)
        # снять значения столбца, удалить, вставить на новое место
        col_values = [ws.cell(row=r, column=from_index).value for r in range(1, ws.max_row + 1)]
        ws.delete_cols(from_index, 1)
        ws.insert_cols(to_index, 1)
        for r, val in enumerate(col_values, start=1):
            ws.cell(row=r, column=to_index, value=val)
        self._save(wb, path)
        return {"path": path, "sheet": sheet, "column": column, "index": to_index,
                "broken_formulas": broken}

    # ═══ ФОРМУЛЫ / ФОРМАТ / ВАЛИДАЦИЯ ═══

    def insert_formula(self, path: str, sheet: str, cell: str, formula: str, overwrite: bool = False) -> dict:
        """Формула в ячейку. Защита: не перезаписывает существующую формулу молча."""
        wb = self._load(path)
        ws = self._sheet(wb, sheet)
        target = ws[cell]
        existing = target.value
        if isinstance(existing, str) and existing.startswith("=") and not overwrite:
            raise ExcelError(
                "FORMULA_PROTECTED", f"В ячейке {cell} уже есть формула: {existing}",
                reason="Перезапись критической формулы запрещена молча. Передай overwrite=true осознанно.",
            )
        target.value = formula if formula.startswith("=") else f"={formula}"
        self._save(wb, path)
        return {"path": path, "sheet": sheet, "cell": cell, "formula": target.value}

    def apply_formatting(self, path: str, sheet: str, target: str,
                         fill: str | None = None, bold: bool | None = None,
                         font_color: str | None = None) -> dict:
        """Стили на ячейку/диапазон (A1 или A1:C3). fill/font_color — HEX 'RRGGBB'."""
        from openpyxl.styles import PatternFill, Font
        wb = self._load(path)
        ws = self._sheet(wb, sheet)
        cells = ws[target]
        # нормализуем в плоский список ячеек (одна ячейка → кортеж кортежей у диапазона)
        flat: list = []
        if hasattr(cells, "__iter__"):
            for row in cells:
                flat.extend(row if hasattr(row, "__iter__") else [row])
        else:
            flat = [cells]
        for c in flat:
            if fill:
                c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            if bold is not None or font_color:
                c.font = Font(bold=bool(bold), color=font_color or None)
        self._save(wb, path)
        return {"path": path, "sheet": sheet, "target": target, "cells": len(flat)}

    def set_validation(self, path: str, sheet: str, column: str, allowed: list[str]) -> dict:
        """Выпадающий список (Data Validation) на весь столбец — материализует enum."""
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        wb = self._load(path)
        ws = self._sheet(wb, sheet)
        headers = self._headers(ws)
        if column not in headers:
            raise ExcelError("COLUMN_NOT_FOUND", f"Столбец '{column}' не найден в листе '{sheet}'.",
                             reason="Сначала добавь столбец через excel_add_column.")
        letter = get_column_letter(headers[column])
        formula = '"' + ",".join(allowed) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        dv.add(f"{letter}2:{letter}1048576")  # со строки 2 (без заголовка)
        ws.add_data_validation(dv)
        self._save(wb, path)
        return {"path": path, "sheet": sheet, "column": column, "allowed": allowed}

    def read_range(self, path: str, sheet: str, cell_range: str) -> dict:
        """ОТЛАДКА: сырой 2D-массив. НЕ рабочий путь чтения (см. json_read_snapshot)."""
        wb = self._load(path)
        ws = self._sheet(wb, sheet)
        matrix = [[c.value for c in row] for row in ws[cell_range]] if cell_range else []
        return {"path": path, "sheet": sheet, "range": cell_range, "values": matrix,
                "note": "Отладочное чтение сырых ячеек. Рабочее чтение данных — json_read_snapshot."}

    _ERROR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!")

    def _recalc_via_lo(self, src: Path) -> Path:
        """Пересчитать формулы реальным движком LibreOffice headless → путь к вычисленной копии.

        openpyxl формулы не считает; LO recalc'ает при загрузке и сохраняет значения.
        Профиль уникален на вызов (без конфликта параллельных soffice). RECALC_UNAVAILABLE — честно.
        """
        import shutil
        import subprocess
        import tempfile
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if not soffice:
            raise ExcelError("RECALC_UNAVAILABLE", "Движок пересчёта (LibreOffice) недоступен.",
                             reason="Установи libreoffice-calc для валидации формул реальным пересчётом.")
        td = Path(tempfile.mkdtemp(prefix="xlsx_recalc_"))
        try:
            subprocess.run(
                [soffice, "--headless", "--calc", f"-env:UserInstallation=file://{td}/profile",
                 "--convert-to", "xlsx:Calc MS Excel 2007 XML", "--outdir", str(td), str(src)],
                check=True, capture_output=True, timeout=120,
            )
        except subprocess.TimeoutExpired:
            shutil.rmtree(td, ignore_errors=True)
            raise ExcelError("RECALC_UNAVAILABLE", "Пересчёт формул превысил таймаут.",
                             reason="Файл слишком большой/сложный или LibreOffice завис.")
        except subprocess.CalledProcessError as e:
            shutil.rmtree(td, ignore_errors=True)
            raise ExcelError("RECALC_UNAVAILABLE", "Пересчёт формул не удался.",
                             reason=f"LibreOffice вернул ошибку: {e.stderr.decode('utf-8', 'ignore')[:200]}")
        out = td / src.name
        if not out.exists():
            shutil.rmtree(td, ignore_errors=True)
            raise ExcelError("RECALC_UNAVAILABLE", "Пересчёт не дал выходного файла.",
                             reason="LibreOffice не создал файл — проверь установку.")
        return out

    def validate_formulas(self, path: str) -> dict:
        """Валидация формул РЕАЛЬНЫМ пересчётом (LibreOffice) — ловит #DIV/0!/#REF! и пр. (F29).

        Раньше был греп токенов по несчитанной книге (openpyxl) = театр: `=1/0` не ловился.
        """
        import shutil
        import openpyxl
        src = self._resolve(path)
        if not src.exists():
            raise ExcelError("WORKBOOK_NOT_FOUND", f"Книга не найдена: {path}",
                             reason="Создай книгу через excel_create_workbook или проверь путь.",
                             suggested_tool="excel_create_workbook")
        recalced = self._recalc_via_lo(src)
        try:
            wb = openpyxl.load_workbook(recalced, data_only=True)
            errors = []
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for c in row:
                        if isinstance(c.value, str) and any(t in c.value for t in self._ERROR_TOKENS):
                            errors.append({"sheet": ws.title, "cell": c.coordinate, "value": c.value})
        finally:
            shutil.rmtree(recalced.parent, ignore_errors=True)
        return {"path": path, "errors": errors, "ok": len(errors) == 0}
